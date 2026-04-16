from __future__ import annotations

import json
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db import BASE_DIR, TEMPLATE_DIR
from app.services.uploads import EXCEL_LOGO_SIZE, prepare_logo_for_excel

MAPPING_PATH = BASE_DIR / "app" / "template_mapping.json"
TEMPLATE_PATH = TEMPLATE_DIR / "quote_template.xlsx"
DEFAULT_FORMULA_ROWS = 17
TOTALS_GAP_ROWS = 5

LINE_SIDE = Side(style="thin", color="7FA3C8")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9ECFA")
TITLE_FILL = PatternFill("solid", fgColor="F7FBFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1780C7")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="0F4E82")
VALUE_FONT = Font(name="Calibri", size=10, color="0B2E4A")
TABLE_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="0F4E82")
BODY_FONT = Font(name="Calibri", size=10, color="000000")


def load_mapping() -> dict:
    return json.loads(MAPPING_PATH.read_text(encoding="utf-8"))


def _apply_calc_mode(workbook) -> None:
    if getattr(workbook, "calculation", None):
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True


def _rounded_formula(expression: str, rounding_mode: str) -> str:
    if rounding_mode == "2dec":
        return f"ROUND({expression},2)"
    if rounding_mode == "nearest10":
        return f"ROUND(({expression})/10,0)*10"
    if rounding_mode == "nearest100":
        return f"ROUND(({expression})/100,0)*100"
    return f"ROUND({expression},0)"


def _line_total_formula(row: int, item_columns: dict[str, str], rounding_mode: str) -> str:
    qty_ref = f"{item_columns['qty']}{row}"
    price_ref = f"{item_columns['price_unit']}{row}"
    expression = _rounded_formula(f"{qty_ref}*{price_ref}", rounding_mode)
    return f'=IF(OR({qty_ref}="",{price_ref}=""),"",{expression})'


def _sales_subtotal_formula(start_row: int, end_row: int, item_columns: dict[str, str]) -> str:
    return f"SUM({item_columns['line_total']}{start_row}:{item_columns['line_total']}{end_row})"


def _apply_border(cell) -> None:
    cell.border = Border(left=LINE_SIDE, right=LINE_SIDE, top=LINE_SIDE, bottom=LINE_SIDE)


def _configure_print_layout(sheet, last_row: int) -> None:
    sheet.freeze_panes = None
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape"
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.page_margins.header = 0.15
    sheet.page_margins.footer = 0.15
    sheet.print_area = f"A1:F{last_row}"


def _style_table_row(sheet, row: int) -> None:
    for column in range(1, 7):
        cell = sheet.cell(row=row, column=column)
        _apply_border(cell)
        cell.font = copy(BODY_FONT)
        horizontal = "left"
        if column == 3:
            horizontal = "center"
        if column in (4, 5, 6):
            horizontal = "right"
        cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=column == 2)
        if column == 4:
            cell.number_format = "0.00"
        elif column in (5, 6):
            cell.number_format = "#,##0"


def ensure_template() -> Path:
    mapping = load_mapping()
    workbook = Workbook()
    _apply_calc_mode(workbook)

    sheet = workbook.active
    sheet.title = mapping["sheet_name"]
    _configure_print_layout(sheet, 40)

    column_widths = [17.5, 42.5, 11.5, 11.5, 15, 17]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in range(1, 40):
        sheet.row_dimensions[row].height = 21
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[7].height = 24

    sheet.merge_cells("A1:B5")
    for row in range(1, 6):
        for column in range(1, 3):
            cell = sheet.cell(row=row, column=column)
            cell.fill = WHITE_FILL
            _apply_border(cell)
    sheet["A1"] = None

    sheet.merge_cells("C1:F1")
    sheet["C1"] = "COTIZACION EXPLORATORIA"
    sheet["C1"].fill = TITLE_FILL
    sheet["C1"].font = TITLE_FONT
    sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
    _apply_border(sheet["C1"])

    header_labels = {
        "C2": "Ubicacion",
        "C3": "Cliente",
        "C4": "Fecha",
        "E4": "Solicitado por",
        "C5": "Cotizacion",
        "E5": "Moneda",
    }
    for address, value in header_labels.items():
        cell = sheet[address]
        cell.value = value
        cell.fill = HEADER_FILL
        cell.font = LABEL_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        _apply_border(cell)

    header_values = ("D2", "D3", "D4", "F4", "D5", "F5")
    for address in header_values:
        cell = sheet[address]
        cell.fill = WHITE_FILL
        cell.font = VALUE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        _apply_border(cell)

    for merged in ("D2:F2", "D3:F3"):
        sheet.merge_cells(merged)
        _apply_border(sheet[merged.split(":")[0]])

    for column in range(1, 7):
        cell = sheet.cell(row=6, column=column)
        cell.fill = WHITE_FILL
        _apply_border(cell)

    headers = ["Item", "Descripcion", "Unidad", "Cantidad", "Vr Material", "Vr Total"]
    for index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=7, column=index, value=value)
        cell.fill = HEADER_FILL
        cell.font = TABLE_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(cell)

    formula_end_row = mapping["items"]["start_row"] + DEFAULT_FORMULA_ROWS - 1
    for row in range(mapping["items"]["start_row"], formula_end_row + 1):
        _style_table_row(sheet, row)

    workbook.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, 7):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def _ensure_item_rows(sheet, start_row: int, count: int) -> None:
    if count <= DEFAULT_FORMULA_ROWS:
        return
    for row in range(start_row + DEFAULT_FORMULA_ROWS, start_row + count):
        _copy_row_style(sheet, start_row, row)


def build_quote_workbook(quote: dict, settings: dict) -> BytesIO:
    mapping = load_mapping()
    template_path = ensure_template()
    try:
        workbook = load_workbook(template_path)
    except (FileNotFoundError, BadZipFile):
        template_path = ensure_template()
        workbook = load_workbook(template_path)
    _apply_calc_mode(workbook)
    sheet = workbook[mapping["sheet_name"]]

    sheet[mapping["cells"]["title"]] = quote["title"]
    sheet[mapping["cells"]["location"]] = quote["location"]
    sheet[mapping["cells"]["client"]] = quote["client_name"]
    sheet[mapping["cells"]["date"]] = datetime.fromisoformat(quote["quote_date"]).strftime("%d/%m/%Y")
    sheet[mapping["cells"]["requested_by"]] = quote["requested_by"]
    sheet[mapping["cells"]["quote_number"]] = quote["quote_number"]
    sheet[mapping["cells"]["currency"]] = quote["currency_code"]

    logo_image = prepare_logo_for_excel(settings.get("logo_filename"), EXCEL_LOGO_SIZE)
    if logo_image:
        logo_buffer = BytesIO()
        logo_image.save(logo_buffer, format="PNG")
        logo_buffer.seek(0)
        image = ExcelImage(logo_buffer)
        image._source_buffer = logo_buffer
        image.width, image.height = EXCEL_LOGO_SIZE
        sheet.add_image(image, mapping["ranges"]["logo"].split(":")[0])

    start_row = mapping["items"]["start_row"]
    items = quote["items"] or []
    formula_row_count = max(DEFAULT_FORMULA_ROWS, len(items) + 2)
    _ensure_item_rows(sheet, start_row, formula_row_count)

    item_columns = mapping["items"]["columns"]
    formula_end_row = start_row + formula_row_count - 1
    number_format = "#,##0.00" if settings.get("rounding_mode") == "2dec" else "#,##0"

    for row in range(start_row, formula_end_row + 1):
        sheet[f"{item_columns['sku']}{row}"] = None
        sheet[f"{item_columns['description']}{row}"] = None
        sheet[f"{item_columns['unit']}{row}"] = None
        sheet[f"{item_columns['qty']}{row}"] = None
        sheet[f"{item_columns['price_unit']}{row}"] = None
        sheet[f"{item_columns['line_total']}{row}"] = _line_total_formula(row, item_columns, settings.get("rounding_mode", "integer"))
        _style_table_row(sheet, row)
        sheet[f"{item_columns['price_unit']}{row}"].number_format = number_format
        sheet[f"{item_columns['line_total']}{row}"].number_format = number_format

    for index, item in enumerate(items):
        row = start_row + index
        sheet[f"{item_columns['sku']}{row}"] = item["sku"]
        sheet[f"{item_columns['description']}{row}"] = item["description"]
        sheet[f"{item_columns['unit']}{row}"] = item["unit"]
        sheet[f"{item_columns['qty']}{row}"] = float(item["qty"])
        sheet[f"{item_columns['price_unit']}{row}"] = float(item["price_unit"])

    totals_row = formula_end_row + TOTALS_GAP_ROWS
    for gap_row in range(formula_end_row + 1, totals_row):
        for column in range(1, 7):
            cell = sheet.cell(row=gap_row, column=column)
            _apply_border(cell)
            cell.fill = WHITE_FILL

    subtotal_ref = f"{mapping['totals']['value_column']}{totals_row}"
    tax_ref = f"{mapping['totals']['value_column']}{totals_row + 1}"
    total_ref = f"{mapping['totals']['value_column']}{totals_row + 2}"
    sales_subtotal_formula = _sales_subtotal_formula(start_row, formula_end_row, item_columns)

    for offset, label in enumerate(("Total costos directos", "IVA", "Total")):
        row = totals_row + offset
        label_cell = f"{mapping['totals']['label_column']}{row}"
        value_cell = f"{mapping['totals']['value_column']}{row}"
        sheet[label_cell] = label
        sheet[label_cell].fill = HEADER_FILL
        sheet[value_cell].fill = HEADER_FILL
        sheet[label_cell].font = LABEL_FONT
        sheet[value_cell].font = LABEL_FONT
        sheet[label_cell].alignment = Alignment(horizontal="left", vertical="center")
        sheet[value_cell].alignment = Alignment(horizontal="right", vertical="center")
        _apply_border(sheet[label_cell])
        _apply_border(sheet[value_cell])
        sheet[value_cell].number_format = number_format

    tax_factor = float(quote["tax_rate"]) / 100
    sheet[subtotal_ref] = f"={sales_subtotal_formula}"
    sheet[tax_ref] = f"={_rounded_formula(f'({sales_subtotal_formula})*{tax_factor}', settings.get('rounding_mode', 'integer'))}"
    sheet[total_ref] = f"={_rounded_formula(f'({sales_subtotal_formula})+{tax_ref}', settings.get('rounding_mode', 'integer'))}"
    _configure_print_layout(sheet, totals_row + 2)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
